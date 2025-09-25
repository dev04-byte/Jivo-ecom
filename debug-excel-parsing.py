import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\singh\Downloads\Purchase Order PO-1357102 (2).xlsx')
ws = wb.active

print("=== COMPLETE EXCEL DATA ANALYSIS ===")
print("=" * 80)

# Find header row
header_row = 13
print(f"\nHeader Row {header_row}:")
headers = []
for col in range(1, ws.max_column + 1):
    value = ws.cell(row=header_row, column=col).value
    headers.append(value)
    if value:
        print(f"  Column {col} (index {col-1}): {value}")

print(f"\nData Rows Analysis:")
print("-" * 80)

# Analyze each data row
for row_num in range(14, 19):  # Rows 14-18 are data rows
    print(f"\nRow {row_num}:")
    row_data = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row_num, column=col).value
        row_data.append(value)

    # Print only non-empty values with their positions
    for i, value in enumerate(row_data):
        if value is not None and str(value).strip():
            print(f"  Column {i+1} (index {i}): {value}")

print(f"\n=== CURRENT PARSER MAPPING (needs fixing) ===")
print(f"Current parser expects:")
print(f"  sNo = row[0]           // Should be: {ws.cell(row=14, column=1).value}")
print(f"  articleId = row[1]     // Should be: {ws.cell(row=14, column=2).value}")
print(f"  articleName = row[5]   // Should be: {ws.cell(row=14, column=6).value}")
print(f"  hsnCode = row[8]       // Should be: {ws.cell(row=14, column=9).value}")
print(f"  mrp = row[11]          // Should be: {ws.cell(row=14, column=12).value}")
print(f"  baseCostPrice = row[13] // Should be: {ws.cell(row=14, column=14).value}")
print(f"  quantity = row[15]     // Should be: {ws.cell(row=14, column=16).value}")
print(f"  baseAmount = row[16]   // Should be: {ws.cell(row=14, column=17).value}")
print(f"  igstCess = row[18]     // Should be: {ws.cell(row=14, column=19).value}")
print(f"  igstCessAmount = row[19] // Should be: {ws.cell(row=14, column=20).value}")
print(f"  total = row[21]        // Should be: {ws.cell(row=14, column=22).value}")

print(f"\n=== CORRECT MAPPING NEEDED ===")
print("Based on the actual Excel structure:")

# Find correct column indices for each field
data_row = 14
for col in range(1, ws.max_column + 1):
    value = ws.cell(row=data_row, column=col).value
    header = ws.cell(row=header_row, column=col).value
    if value is not None and str(value).strip():
        print(f"  Column {col} (index {col-1}): Header='{header}' Value='{value}'")