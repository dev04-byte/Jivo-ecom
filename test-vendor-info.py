import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\singh\Downloads\Purchase Order PO-1357102 (2).xlsx')
ws = wb.active

print("Vendor Information Analysis:")
print("=" * 80)

# Check the first 12 rows for vendor info
for row_idx in range(1, 13):
    row_values = []
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_values.append(f"Col{col_idx}: {str(cell.value)[:40]}")
    if row_values:
        print(f"Row {row_idx}: {' | '.join(row_values)}")

print("\n\nExtracted Vendor Information:")
print("-" * 80)

# Based on the structure, extract vendor info
vendor_name = ws.cell(row=6, column=2).value if ws.cell(row=6, column=2).value else ""
vendor_code = ws.cell(row=7, column=2).value if ws.cell(row=7, column=2).value else ""
vendor_gst = ws.cell(row=8, column=2).value if ws.cell(row=8, column=2).value else ""

print(f"Vendor Name (Row 6, Col 2): {vendor_name}")
print(f"Vendor Code (Row 7, Col 2): {vendor_code}")
print(f"Vendor GST (Row 8, Col 2): {vendor_gst}")

# Check for PO info in merged cells
print("\n\nPO Information from merged cell (Row 1, Col 18):")
po_info = ws.cell(row=1, column=18).value
if po_info:
    print(po_info)