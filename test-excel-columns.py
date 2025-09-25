import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\singh\Downloads\Purchase Order PO-1357102 (2).xlsx')
ws = wb.active

# Find the header row (row 13 based on our previous analysis)
header_row = 13
data_row = 14  # First data row

print("Column Mapping Analysis:")
print("=" * 80)

# Get headers
headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    if cell.value:
        headers.append((col-1, cell.value))  # Store 0-based index and header name

print("\nHeaders with 0-based indices (for array access):")
for idx, header in headers:
    print(f"  Index {idx}: {header}")

print("\n\nFirst Data Row Values:")
print("-" * 80)
for idx, header in headers:
    value = ws.cell(row=data_row, column=idx+1).value  # +1 for 1-based Excel columns
    print(f"  {header}: {value}")

print("\n\nCorrect Column Mapping for Parser:")
print("-" * 80)
print("const sNo = row[0];           // S.No")
print("const articleId = row[1];     // Article Id")
print("const articleName = row[2];   // Article Name")
print("const hsnCode = row[3];       // HSN Code")
print("const mrp = row[4];           // MRP (₹)")
print("const baseCostPrice = row[5]; // Base Cost Price (₹)")
print("const quantity = row[6];      // Quantity")
print("const baseAmount = row[7];    // Base Amount (₹)")
print("const igstCess = row[8];      // IGST (%) / cess (%)")
print("const igstCessAmount = row[9]; // IGST (₹) / cess")
print("const total = row[10];        // Total Amount (₹)")