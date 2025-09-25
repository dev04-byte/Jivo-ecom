import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\singh\Downloads\Purchase Order PO-1357102 (2).xlsx')
ws = wb.active

print("Sheet name:", ws.title)
print("Total rows:", ws.max_row)
print("Total columns:", ws.max_column)
print("\nFirst 20 rows:")
print("-" * 80)

for i in range(1, min(21, ws.max_row + 1)):
    row = []
    for j in range(1, min(10, ws.max_column + 1)):
        cell_value = ws.cell(row=i, column=j).value
        if cell_value:
            row.append(str(cell_value)[:30])  # Truncate long values
    if row:  # Only print non-empty rows
        print(f"Row {i}: {row}")

print("\nLooking for header row with 'S.No' or similar...")
for i in range(1, min(30, ws.max_row + 1)):
    first_cell = ws.cell(row=i, column=1).value
    if first_cell and ('S.No' in str(first_cell) or 'S No' in str(first_cell) or 'Article' in str(first_cell)):
        print(f"\nFound potential header at row {i}:")
        header = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column + 1)]
        print([h for h in header if h])